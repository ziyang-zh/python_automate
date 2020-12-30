import openpyxl

n=3
m=2
readwb=openpyxl.load_workbook('blank.xlsx')
writewb=openpyxl.Workbook()
readsheet=readwb['Sheet']
writesheet=writewb['Sheet']

for i in range(1,readsheet.max_row+1):
	for j in range(1,readsheet.max_column+1):
		writesheet.cell(row=j,column=i).value=readsheet.cell(row=i,column=j).value

writewb.save('transposition.xlsx')