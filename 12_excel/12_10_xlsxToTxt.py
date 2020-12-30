import openpyxl


wb=openpyxl.load_workbook('fruit.xlsx')
sheet=wb['Sheet']
with open('fruit.txt','w') as file:
	for i in range(1,sheet.max_row+1):
		file.write(sheet.cell(row=i,column=1).value+'\n')

