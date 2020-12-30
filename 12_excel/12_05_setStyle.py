import openpyxl

wb=openpyxl.Workbook()
sheet=wb['Sheet']
italic24Font=openpyxl.styles.Font(size=24,italic=True)
sheet['A1'].font=italic24Font
sheet['A1']='Hello world!'

fontObj1=openpyxl.styles.Font(name='Times New Roman',bold=True)
sheet['A2'].font=fontObj1
sheet['A2']='Bold Times New Roman'

fontObj2=openpyxl.styles.Font(size=24,italic=True)
sheet['B3'].font=fontObj2
sheet['B3']='24 pt Italic'

sheet['A5']=200
sheet['A6']=300
sheet['A7']=400
sheet['A8']='=SUM(A5:A7)'
print(sheet['A8'].value)

sheet['A10']='Tall row'
sheet['B11']='Wide column'
sheet.row_dimensions[10].height=70
sheet.column_dimensions['B'].width=20

sheet.merge_cells('C1:E3')
sheet['C1']='Twelve cells merged together.'
sheet.merge_cells('D5:E5')
sheet['D5']='Two merged cells.'
sheet.unmerge_cells('D5:E5')

sheet.freeze_panes='A2'

for i in range(1,11):
	sheet['A'+str(i+10)]=i+10
refObj=openpyxl.chart.Reference(sheet,min_row=11,min_col=1,max_row=20)

seriesObj=openpyxl.chart.Series(refObj,title='First series')
chartObj=openpyxl.chart.BarChart()
chartObj.title='My Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj,'C5')

wb.save('styled.xlsx')

