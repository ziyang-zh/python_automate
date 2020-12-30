import openpyxl

n=3
m=2
readwb=openpyxl.load_workbook('multiplication.xlsx')
writewb=openpyxl.Workbook()
readsheet=readwb['Sheet']
writesheet=writewb['Sheet']

for i in range(1,n):
	for j in range(1,readsheet.max_column+1):
		writesheet.cell(row=i,column=j).value=readsheet.cell(row=i,column=j).value

for i in range(n,readsheet.max_row+1):
	for j in range(1,readsheet.max_column+1):
		writesheet.cell(row=i+m,column=j).value=readsheet.cell(row=i,column=j).value

writewb.save('blank.xlsx')