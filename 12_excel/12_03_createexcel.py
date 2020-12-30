import openpyxl

wb=openpyxl.Workbook()
print(wb.sheetnames)
sheet=wb.active
print(sheet.title)
sheet.title='Spam Bacon Eggs Sheet'
print(wb.sheetnames)

wb.create_sheet()
print(wb.sheetnames)
wb.create_sheet(index=0,title='First Sheet')
print(wb.sheetnames)
wb.create_sheet(index=2,title='Middle Sheet')
print(wb.sheetnames)

wb.remove(wb['Middle Sheet'])
wb.remove(wb['Spam Bacon Eggs Sheet'])
print(wb.sheetnames)

sheet=wb['Sheet']
sheet['A1']='Hello world!'
print(sheet['A1'].value)

wb.save('create.xlsx')