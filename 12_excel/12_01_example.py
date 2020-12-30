import openpyxl

wb=openpyxl.load_workbook('example.xlsx')
print(type(wb))
print(wb.sheetnames)
sheet=wb['Sheet3']
print(sheet)
print(sheet.title)
sheet=wb.active
print(sheet)

print(sheet['A1'])
print(sheet['A1'].value)
c=sheet['B1']
print(c.value)
print('Row '+str(c.row)+', Column '+str(c.column)+' is '+c.value)
print('Cell '+c.coordinate+' is '+c.value)
print(sheet['C1'].value)

print(sheet.cell(row=1,column=2))
print(sheet.cell(row=1,column=2).value)
for i in range(1,8,2):
	print(i,sheet.cell(row=i,column=2).value)

print(sheet.max_row)
print(sheet.max_column)

print(openpyxl.utils.get_column_letter(1))
print(openpyxl.utils.get_column_letter(2))
print(openpyxl.utils.get_column_letter(27))
print(openpyxl.utils.get_column_letter(900))
print(openpyxl.utils.get_column_letter(sheet.max_column))
print(openpyxl.utils.column_index_from_string('A'))
print(openpyxl.utils.column_index_from_string('AA'))

print(sheet['A1':'C3'])
for rowOfCellObjects in sheet['A1':'C3']:
	for cellObj in rowOfCellObjects:
		print(cellObj.coordinate,cellObj.value)
	print('--- END OF ROW ---')

print(list(sheet.columns)[1])
for cellObj in list(sheet.columns)[1]:
	print(cellObj.value)