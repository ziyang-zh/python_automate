import openpyxl

n=6
wb=openpyxl.Workbook()
sheet=wb['Sheet']
boldFont=openpyxl.styles.Font(bold=True)

for i in range(1,n+1):
	sheet.cell(row=1,column=i+1).font=boldFont
	sheet.cell(row=1,column=i+1).value=i
	sheet.cell(row=i+1,column=1).font=boldFont
	sheet.cell(row=i+1,column=1).value=i

for i in range(1,n+1):
	for j in range(1,n+1):
		sheet.cell(row=i+1,column=j+1).value=i*j

wb.save('multiplication.xlsx')