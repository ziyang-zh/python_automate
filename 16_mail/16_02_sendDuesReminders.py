import openpyxl,smtplib,sys
from email.header import Header
from email.mime.text import MIMEText

wb=openpyxl.load_workbook('duesRecords.xlsx')
sheet=wb['Sheet1']

lastCol=sheet.max_column
latestMonth=sheet.cell(row=1,column=lastCol).value

unpaidMembers={}
for r in range(2,sheet.max_row+1):
	payment=sheet.cell(row=r,column=lastCol).value
	if payment!='paid':
		name=sheet.cell(row=r,column=1).value
		email=sheet.cell(row=r,column=2).value
		unpaidMembers[name]=email

smtpObj=smtplib.SMTP('smtp.qq.com',587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('my_address@qq.com','my_authorization')

for name,email in unpaidMembers.items():
	mailsubject=latestMonth+" dues unpaid."
	mailtext="Dear "+name+", \nRecords show that you have not paid dues for "+latestMonth+". Please make this payment as soon as possible. Thank you!"

	msg=MIMEText(mailtext,'html','utf-8')
	msg['Subject']=Header(mailsubject,'utf-8')

	print('Sending email to '+email+'...')
	sendmailStatus=smtpObj.sendmail('my_address@qq.com',email,msg.as_string())
	if sendmailStatus!={}:
		print('There was a problem sending to '+email+': '+sendmailStatus)
smtpObj.quit()






